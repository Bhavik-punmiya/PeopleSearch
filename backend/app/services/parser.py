import io
import csv
from openpyxl import load_workbook

def get_file_preview(file_content: bytes, filename: str):
    rows = []
    if filename.lower().endswith('.csv'):
        try:
            content = file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file_content.decode('latin-1')
        reader = csv.DictReader(io.StringIO(content))
        headers = reader.fieldnames or []
        rows = [dict(r) for r in list(reader)[:5]]
    elif filename.lower().endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active # Preview just the first sheet
        headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
        for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))
    else:
        raise ValueError("Unsupported file format")
        
    return {"headers": headers, "samples": rows}

def parse_hr_file_with_mapping(file_content: bytes, filename: str, mapping: dict):
    rows = []
    if filename.lower().endswith('.csv'):
        try:
            content = file_content.decode('utf-8-sig')
        except UnicodeDecodeError:
            content = file_content.decode('latin-1')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif filename.lower().endswith('.xlsx'):
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    rows.append(dict(zip(headers, row)))
            
    parsed_people = []
    for row in rows:
        # Use provided mapping
        name_col = mapping.get('name')
        comp_col = mapping.get('company')
        email_col = mapping.get('email')
        link_col = mapping.get('linkedin')

        name = str(row.get(name_col, "")) if name_col else ""
        company = str(row.get(comp_col, "")) if comp_col else ""
        email = str(row.get(email_col, "")) if email_col else ""
        linkedin = str(row.get(link_col, "")) if link_col else ""
        
        if not email and not linkedin:
            continue
            
        parsed_people.append({
            "company_name": company,
            "person_name": name,
            "email_id": email or linkedin or "",
            "source_file": filename,
            "raw_data": {str(k): str(v) for k, v in row.items() if k is not None}
        })
    return parsed_people
